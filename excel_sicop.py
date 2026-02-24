# ============================================================================
# GENERADOR DE EXCEL SICOP - FORMATO INSTITUCIONAL
# ============================================================================

import io
import base64
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

from config import (
    formatear_fecha, obtener_ultimo_dia_habil
)

# Logo SADER en Base64
LOGO_BASE64 = """iVBORw0KGgoAAAANSUhEUgAAAcYAAABbCAMAAADeM9UyAAAAAXNSR0IArs4c6QAAAAlwSFlzAAAXEQAAFxEByibzPwAAAIRQTFRFR3BMYT40oYMsSR0xoYMuRhwwooQuQx0uoYQtQR4uooQtRhwvooQtQh0uooQtShsxooMtRxwwooQtooMtWhY2JSUlooQtWxY2JSUlooMtooQtWhY2JSUlooQtWxY2JSUlooQtWxY2JSUlWxY3JSUlpIUuWxY3JSUlWxY3Xxg5Wxc3JycnIa3V3QAAACl0Uk5TAAMJCxIWHiMqMjY/Q0pQWV5ma3d4eIONj5GgpKqyusbJ0drl6ery9Pnr4LqIAAAaIElEQVR42uzXgWrrOBAF0CtZ0VMmymQyUxSk4pBSGuL8/wfumqSFR2Gb5bnwXHw+wJh7mfEYi8VisVgsFovFYrFYfC/nHQDvlyTmyodIsRRL4FKSI45LJnPjPJilj5VzoxaoUrUCYBnLGfGRGicJLRRBbRWxp54TgFTCEs88BM8W60kVLZmi1orUe+kDAH4LyMt2nQFXzUlK1p+SsTapqXIxCX0GvJ2CvuXo8XfoMJ3VZv+06/BjcPOJiUNTrWISEYWDWAkAWZNTQa0Jf4Ht8/F57zCJ9eF4GYZhj5/ABQBoAk1GPoXocedjAGLR3lrilkvC/+FW2/3GYVLd4Tz86+CmKPH5dRiu1+tw7DB/nksAQD1bpIBPXMpcaugrU7SEh3VPr9fr5YApucMt+PN6grEeSxwNLyvMX66384UL4UPknFPAnVOrxozUc3x8bsaUhssGE9pe78lv8ccOw/Xm8gNqDDHUvjf3+9WaamZhMX0/bGJVCo6b1YjH7O+BP2NCx/foN/hj65fxYZ+X6m7fYW6itpKIGuGDC8rSKAsnomySPUbcJPcteSI8YnVPfHhdYzr35IfjaorRPn+uEW5znuHJYxRqI7WIGwcPskxcNZGoFiKxjFEouQpCK3jE5nK9GQ6TT+PwusEUDsPnGrvjMBwwN6ZAFU64CUQleTUWsV64JFNm0moEwKVTCs1i9Pja0xjR5AfEbrxTLy9bTGJ9ve/9Dh+2l/nVmHLuxQsDIQJAEiMTSjEXJulZJbMa53YyD8CzNMutqcdXunFubobd1H+NK0zj1/lTjeN7z61GstpKKxIQRQFE07G+UksOxlyMKrGwZOYkEQC05pN4UYcvrC/Xd8PRYTIOo6lrdL8tkeEJc5L6DG6JPLiqBsQqpdTExRrnrJQaF1IVzcTKmgAgVAFE3UOfnftAXtaYzvfWuL98T43OdZ3D94i1z9AM0FuVFMKpF65CapVJVTPXrHksVUnE2AgAjIGAr6zGk/Jy/u8j59d6u9vvNutfKweMun/Yudo1N3GeLRO+McbYjKhMncKySWF6/uf3hi8nDNNpO0+ud//03p3diQEjdCNZkjWJkhP8GqfodBjJq7quqzxibwWJkuDNuUEUsfdpDOrX8f1VgAXR/aZO2Ah+B0FSNudzdz7XhbvguRBW6xBAtUalLDVCSImCoxYSUQitlSKutAyllAqX8kCsYhZylcKHKKa3ui6u40+DnFPRdK/jhKE/N3VVllXd9K8b40Geb7otqtvh+wxJde6a6M1U13FB3+QPVe+iqs/99Zzch/Kybrq+r9g7NAZlt6W6N2Gqqq6LeTzKy6rprtfmLkM0zdxdu+w+kmfuWN3UZeCEK8+v44auejqRXhp6EFs1sYlChCGiNi0JgZhqraRRSpBGThInk1SaJMWzW0VNbQwfgTXj69ifoFn0MpTHfOT2cC6WfUDHlpJ1N1xnRoNqIbtLNsb66WP1Zqp1ktt/N22zsru+LmONqw52K919cKQx6w4C9dFkod11mWYzUhY1/bDOzJaRrLnJW81T1ou855WvoJlmvf90OTwXDFtjcA46Y4yFUEoaUq0USiLpG2KDmqsYDaUoETWSIh+AcR9/wSIk/RLvZYPT0x5FPy76Gu66c+FQMhnXUsbLO1f3PE1vdjedvy8NlddZP33XdZML2DQYOFbG88a3u9kQHWhkLrR2GPvElf8mVG7ddzOzmcTzJO98duHkPZ/uAfvYN1VVd/P4tXwui+Arkst6x8hI5EoKqbgvtTGatOCKrJImRRXPG5CEhIiT3NwKABH/Imm8JttTHCvZ2cziODRVWT2qbxzLJTdc84CJog0lQLXpaLjTWM3c9VVygqA4jxsrLre/c86Cbhu6Bkcam8m+nByOiuAuX+FSy21ovlM5rKfXQTXcp8jv2XOTrEXmWfYCngiufZ/SNYAPpZDWGq1DSK0gHUuFUpGdPCtZjohGGUOEhgOA0h7/0K2eunFNxqpV79UhAJpfTLbEFSuGoatOjyXr800rj2RU7kO9W4THcwLsrqnBae3NzbN+M+0jjRCV9XmNyfquvw7DcC3mkzby70X+aidG416O7vVB3noxxs0w3ZmTq34ekGLwKAQHQdoqCVxJqQQaFEhcm1SjVoREM4+GjAfgoU6tZj78DPkkbbl419cZXXCs8AyVe7plAY2i6SyW9Hc/Ow79sHkp5izD0ciifnSs3Cu55fa7syJH+vp+nI403pkfqyAIIheIZtfNhA9VvHIpzW487+Rttlfg7oqi67Ork7Ehg5yzxyHZEo+VRm20RsI0RqM5yVjTzKM2BmkOiVJPWq4FHOF4GbtoVkzjVLkz1nW1WxD1+zgo751n6oqo6FYPt6plp4l5+iF/s6+yHi3G/W7IsgA63o40QrkcLg6P4+R1a//DicXVyXu+yduPdxr70VU/FuN0Ez0HsfJibTkwDxZ4ClUrmVAaSZCyKDUnJXD6BQ0qREIyRDacrr4Q5+9b4yZ88/PtqsRR4eLanTKdW13ehWx7h505bZEqS67Oad3Ncxxr50NdPONmdnO9R2N1oHEb3d3Grfm549ltvMxr+2apUbdvEalHV9h/DgRnDGLPQy3CxTqVaBXzlVRKkw6V0FZIQagkaoVGoZQtCatIAQAXQsmYfRDgDPmd0rcLQrHGLDsTcq7RuSn3vM3ooqRgPVJtajksvHXfd5mzmbe3rn6Lxhx2KDcTPlh1tsq77clF7uDyO8ubpjod9gsKeBJ8TJXVPvhkCLkPoZW8VQASSWEqpDBkjBTUKklI2nJFIhSCa9NeUgAmLGmKf14Vd46DNccgp3pLY+HC+73Pau6JwlCtwZGj0X0adlpnywrrrHFffCg/QaPj/uhnV4mz3TOyfEpfHVXsOFP5NBrJtBIFQCwUoQJE3koAUDhHNFZZ0oJLpQVKVBp5GnOthY01GgUQog8gxM93Y+/PXB62Zreh6lc0utCwPJ9XrZx2NObD/jqnNyeIu/OexupA486bv6Gxdhd9SKMThFVds5viONOzkAo/1fFSzVE6NWJmkUlFShlttdZaKRQpajRpKIRRAo00CkMZAoQ6TZWJ4SeyOnfjvOrjCOSr8t4ot4s2HjZv6PQfbLa9p7Ea3QL6Dlh+2IZiv0WjE9WNHmzo/B6N52An7x7sFERl//pcGgX3QEpvjXa0JFJrTwcpRDKKyEiDJLWMU0kiJUqlEdooTny6hiylPw1wdo7svAWX7E3c3Uf7EOfs7azxENEdrbHeOc0DisOaBtWvaHRe4Gh6xS9pZHCAq+42XT/5jufSqFPFhV4plSq1JPxZo2iNVsZYnWpNxqo01BhqCkMjPVRhq0KJE93Sj+m4e+xcZl8mK4puFb4P3uqlhi0x273tLnPcdzodrfHUfGyNu4Xw8zSyVdz8VzQ2P3EKa1XWVYmetjZ6hEZzBRN8qbnfKsX9MAyZIE6tEpQqJCKdcj9WimtKYyNCirkNY8MAhI5bxKNXdUHc0K8YHqtpjrflnDmIY1k3HtKyn3UBBzsazx/3rFZPoNHd5pM0snwpuN5+huv1yTT6UqGxEsDzwEelAG0ap1KT9n0tpZBKIxotdYwKWIxakUCUGlNCZn2AVFuKMYQDNorcLsGhCeAxMTyXedFcFzLuWmK/T+NHTSLsz2lsFlH2NAZbdvEJGoNmGOd06dxURZ41bpV9HrwQQOnpR6rYWCW4SFGFSJIMTiCDnDQAhNbGZFAikTThHNqk0mMfdps5OBqv2aFTx50w9gXAH1qj83Y1g3dR/ymN7H0au8/S6OqBTX7aJxxPRqrj0FNaKF9ojWQ4xgpJG9KotZaSlAyn0y5ct0YikqbUhPBzRIuHHK4Ow76e7Tp9Hwgeh3MG8KfW6F6ZQ635dPocjacPaczBgf0ejafGLR4LymfTyGKhGcBke8CUljF4gibehJLaGlJSk0xRKylSD0DpsDVKIWqUxAB8z9fx+73ia2kxWBHlza605vLyq1v3+3O5J+x3aSz3Fda7rS9X1sdItf6YxnVdf5/G4mHo92gsXvfjz7fGsL1YABCIioGnpQROlmKSREoSSYNcCSSNqdA6BC/0jCZEIj27WSl4q1T6XgVnSQHYIdJbxHd66Yu8am6oq0OHiiu/fESjW4eP5lPPuYwLWB5Lrs2exv5A47GKwzbOql1S9Ts0Lhdeo/vtn0wjl6Y1KYDHJUoAzTWXrTGWpySFkkJoQcoYaUgLScgn5qRuVWu1EQCeMmjJ4jt9je5tPjy0M6/yviXP2JxGvGkby34zb3RLbJfAHdWwppLVwecG3YfW6Fg+vpruLFc++nXeGL19jKB/bqTK2gteLnxOAFELIbTPFAlrNEnUklAqacgYrcikhJqUD54ftty2yvKJRkvGav1uVHF0hs2u05HNVnItTkGSV82562/oznUe7aLdX1qjI2qJkNyuV+3sojzuN+7z72g4RKqPfLC8OrlRV65wI0caT0enckN32vnUZ1ojkmrRg6UeY0yKc46YatQ0GSAqpdEI0lbEAg1qxBSAGaGsvEgApiySQXXQcn9s83T6XNTrlHBtzsP4iM5FAvlCYx/A8QaORsfqPPDalEkUJUXdj65PIru6vf796WN1aP7fhdnDto84RI/ar/fF2juN+a4B8NCqkrnPT6ZRtIYu1l8rATGnVMrYnz6nqJEQtVRSa8MpTRVqQ1qjYKAkN3yiEdLWGmvjd4zxIxpfr/n60TW8OLhmKHfCe23K2Zv2muJ+9Wvfdf0wLu5gz1qz9my45oFq34yx6d9Jdq2iIGuuCzGu92aomKPD0Xjneeijo1O9u3yv7J9dxeEXKy+zLfnGAo+1VQCxnLcQfSLkpHyS3KSpElrNkY1C5aciNnErAfxYttYKcNhZUQNHp+qC1Ucb2rPpzMLFRMfHrR961pyDXnHPQZPD9nOVJVm1qHG/z7W/UbLZy+u1v7o+ylO3jZ6LJMmbq5sn2R/PDwWtVdq6LOupqe/ZTjUkMp5MZxp9KYSVsReTnusyk1WGPsScc01KKK1QE83bVV6IYSsA5EUaLT3YgUXdsQ/Ovc2OKKf7ccIw62s7OhTOZx28lFPYhGHrEHVdIg5jnz+Yg2P4dc1uDmmnW/b2L9xjj0/5Oj6MTtjRuLqE4+JYLNdttx7785OLcb5oEWenyCyJMEylRS516oEDk2RQEXKFChENKYHKD+M2ndNNkhaPyYZzecew5O5Qin5xgnN9KkuyvL4+tkp5iy6P9RnXEvW4V3yqh3HXItll4FC6Q6uvdKGo8x371TFxY84Zru5hd4uy2xIO99xO/OM3FbjLomJ8rlM11ihrZtJUi1yHhL5IbQxhTHpt0PG10qgkSDRKoUTDtQ2BL0tqKJSVsEPi3tlrtNP9Ts3nxXONXRmcgDG3+N/VGTyo8p29hsOrMhefN/vu630P3mLqjuDS0XifbBeGFFdneI9/NNA8TDM0ASxF+THbUz+88UOM1f0k2ipZ9OwqDpOWcNtn8qxV3MZCGOQcW0PCSGB+KKVETEUKQimhteBaTu42BvB5Gmt2iCK3yHM4Bzt6u9fRoS9OzXjcXao3Gtn8djvkh7fbzbPbCjpPjaXX7lxFsEfW9KtLO1cBQN4P1+vQFyvFDl2yfynG67k67Ta9toe7ngsGUM9lxi5aI6cNjkaH7Cba8Dr05zrZHMB4LZ74RY0yjlcm4jCW1ktDHsbcagBhTEqGLkITtUYjhzD1wrm2Cpw8AG3JoucdcvOsKMqyKLLgDb95Ua4oEsiGY90FysfYIyiqekFxenOH20zFDeW+8sPYKcnyPIsYOLhDSVk3TV3mwfwxySYskVRWVnVdVdW+kHS6DZeHZ4Agr+Z5MjZ/yiasV0XlJm/+jgDBTbQsOa1z5zdkJ3gapCXSzqRkS7EUoKQNYzn905IgQ9a2F31pJ8JECH5KCrUP3LbUIsnP/Q1pNbr+jDsaFxw9/W9V2fzDnvHXsIx9Qo7DFAyeBsbbVlJ78WGCD34obItKEQGEPo9T7XG6XChOddsairluSUil5gIetWRvEPAp1OORr3JYQ4S/X973Z+DtRdvWsplToVOAWAnJZctBEI+59i+tECKVfijVRRsbpyblKCRAaNCQscaDT8DtgUTgENXDuAY4f4n5MzC6XJBaNfOILcYxAGgUeLGkwjDW3CCEcRgDGCRtVJoqa/icU14IjTEaPgXXBFzn0bSPlZTnfpxH/rL4Cai2lepi43Tu6RCCZAyauI9CX6y9oFY+yJCLEC2XRqWSg05nDxyb1hhrQ/gkypW1uV/nOv22JmPsLyt/jtjOpRi57iJrVEqhgDma4RpI+J7PJbZWKSkplqm8kA/gWxO21loOnwUr3Xfu3f/fV9FfSj4F2aKlFmGGn2KYylgrz+Oh53ueESq1hrThs/Vhe2ll7AF42AoTxz78D0jqfhhXzFbZFBH8tcXPgTHRGpMuLBpUPATwFFctKVQphYDIzdzpQVJpaSQAeByAGxQp/G8I8ro5dzecz02VB38p/Dw8ScZqIWQI4BNxqVEAeILiWMnUGLykqbah54UxlyoFNrEsQYgULwSfwfGbTQ5fd8JO7I+SOnfR0+FmPp3YH5z9/w4m2hY1XS4CAOJ4dqyaATce00pzIaUHkrhGI9C2YiJeW6GtBj8N3+Ulf3lJTh/a/9uBL9/y/cDLtwQekb3kb3Pu5J8S9ii/5W7yXyAryzL5fWW/fMvg67cIHFhwm+BnZZjk2wv8B4h9Dy94URTDEoNyzUASg5S0SuddZOOlxshQ8olFNDdYHTJw2HHy49/v/0QfeNIvGeyRfSsY7PD1x45X9u3HvwHskXwt3ir7RwnB1wQ+gpv/339/fI3gN/HlJs4/3x9Pz79///eHI3aP7MdX+C/APNUatBc9O1llUaiYGx8YSVIIANxAaoRO154BNMYaQyG8h+jHtyQvGAQvX8oTQFB+eYlOZZJ9iSD/8iWB08uPr2XAii/T4eAlKl6C27+QvHx5WRnIvpRfv2fAituV68j3f76Xy5HsNpbnyZckeMkAkukm0e3C6CW/0ViwL7fJT8ntnGkwK27DLJ/vdEN+Gw5eioXGMr+deoLTJN4yT7TJUOTJSxaUN8nYLMNC47d/Izg5kfIfX7OvP14gm0VPIM9vtzrND/Wf0Qihba1qLziXHTmKUP1fe2fb8rgKhOGx1rU27ykRDYpBLH7w//+/M2PS7lnYPQd2WcqyuUubNurMdK7e9MlTSMZ+vV2g1w/y6GW0X3qtrb7Atb1axLhZ60f2fYw5kdtEyDEbJlwppZMppqKGFGNqZColKh5SLAaagtOkKw0sOZbdw10qKaW6J8WdrElTDACKRqijOLdraHksJfGpLNRWwqgweJBkywHvpoSy7JlYdesC095i+pjwUCQzWGUQCnOnDpZUa3C4mRociDhMxb4wMlMo/I7RAOU1WDqlcxlTiZAS7lfFfe7sxn7WW/+lR477yTa9vSC/dux7aK1Hf95ty6B/9g+/ev3wM4Pva8nFSZiKQTyyK0YoLmMxSiBD7DZbCpoRpOpy5Kpk9CphFEqZMh1OESE3IgeFQQDFYwSXJRjsFM11aAXWlYW5MggFO0ZDGLnBO8MnL4ydoEwpCgCQKTBafmDEYaVKUFhPhxvJqQZMXsPzpuRJKkQ8FHdgFMhHHCV1JSJt9cZIawQoSvVBjPBYbXt/bq3vgUQc58fYj+NMNz2zt2/HdVu99/MFfqTOlSBMiSGVZqlfcjJGBJdzjCXAVAYAZUKoGB0Ac0WxJYRYMbKQBZjcqJJwjzt6tgRkRQymilHiTjJa5gD/xgg131eMzSuT2OENMbAXRgoz1CSGu5IWzo8asByo9oYBIwuEf2CkmDKFWlJOtdzljZFSuRDKJzGyebPrOm72OX916Dg/Ho8Zda9TrqNtAUY/99rqFn4kzpAL8TNN0wl6i8BkDBxEjl3TKGoytXBR6Y2x2uLlxlAkYZRopaZTgDIl483hvA6WilEcGIs4ME7fYpzoTnbbM8UdY5cCTn1hVDlxtBomkcC7UKbmqMHlitEQRgMyh5cbMSWuIow0rLLjdeS1hjuM8Vk30m9U9uG9vcP1xm4tg124Zfv4w1vr76zVt8v1Aj9WF8wSk8SeNN3EmhIn00kkBowaupCf3CSxa1OpGNmB0VEnd2huQR+zkKdmUdXLSQoRshpwZSgNga6NpKmT4x3mCG+MbkLXhCm+3OgoUxKvP3jljrE4k6q3UtcsolkaUxY0YBO+caOMecKqXxhxNr08ML6+a5d0uJGHMiwfdWNV67e5nWf9fKzPdqfH2OUKpNu26m2drYb5fw76VUg5DMCWkJPjbIloQhEcB5Au4XOQIUfVxRyc4yoaxGii5C5Fg8uApuFQUNBQIMLYRAPApjjgpBDo4x8F7V0oYgqcm5RMwJZicBVykAKDuYBMowKgTCZUjASZAcnEGN1AOAJGEE3EWqU4ajBB1fB1OCcjMFA9bqSX0XCAfViEILHY6F5rhpidc6Ci+ey/c/Stf67rc31u3q+3Xh9n4rxf5w21Wu9HuNgR/lNCKVFxKMXrRjImJbWPKUVbQY9ScS4Zl4JWSA5cSRACSFwJJjltldwD8mMvTiKHCcnq6xpR1IkgRQ0DopGMgr13gKRw7DikHY4aJVX0rpbVWr9ZhpUdw0eB8igdUMewUILWYLHvVJzLOvZJMbZuVuv2/hj1tj3RgOtqt9mu3iNDbTePRD93pnDWReNK+OkedeHrYvYbS/+87n3rn3qbtW1Xrb3WYzvPvu/9c7V2RZD2Bp+Tcs4ZBT+P0Uj4K8QArNbPDbHdxrFf9ytw6PYyeo/e1J++xvGF/dq7+3t0gXHtR3+DW7ut8+ZHfLwDjH573M6L4v55+vLU80YOtFt7duPPxThe/ExHIdf2dnbjzxWDKwBcbueF4k+dOnXq1KlTp06dOnXq1KlTp36b/gGN+mEBlT1ZBAAAAABJRU5ErkJggg=="""


def generar_excel_sicop(resultados):
    """
    Genera el archivo Excel de SICOP con formato institucional.
    
    Args:
        resultados: dict con los resultados del procesador SICOP
        
    Returns:
        bytes: contenido del archivo Excel
    """
    metadata = resultados['metadata']
    config = metadata['config']
    resumen = resultados['resumen']
    subtotales = resultados['subtotales']
    congelados = resultados['congelados']
    totales = resultados['totales']
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Edo. Ejercicio UR"
    
    # =========================================================================
    # ESTILOS
    # =========================================================================
    font_header = Font(name='Noto Sans', size=14, bold=True, color='FFFFFF')
    font_title = Font(name='Noto Sans', size=14, bold=True)
    font_data = Font(name='Noto Sans', size=14, bold=True)
    font_subtotal_white = Font(name='Noto Sans', size=14, bold=True, color='FFFFFF')
    font_notes = Font(name='Noto Sans', size=14, bold=True)
    
    # Colores institucionales corregidos
    fill_header = PatternFill(start_color='9B2247', end_color='9B2247', fill_type='solid')  # Vino
    fill_total = PatternFill(start_color='E6D194', end_color='E6D194', fill_type='solid')   # Beige
    fill_subtotal = PatternFill(start_color='002F2A', end_color='002F2A', fill_type='solid') # Verde
    fill_gray = PatternFill(start_color='98989A', end_color='98989A', fill_type='solid')    # Gris
    fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    border_dotted = Border(
        top=Side(style='dotted'),
        bottom=Side(style='dotted'),
        left=Side(style='dotted'),
        right=Side(style='dotted')
    )
    border_none = Border()
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='top')
    
    fmt_money = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'
    fmt_pct = '0.00%'
    
    # =========================================================================
    # ANCHOS DE COLUMNA
    # =========================================================================
    anchos = {'A': 8.71, 'B': 106.57, 'C': 32.57, 'D': 32, 'E': 38, 'F': 35.29, 'G': 32, 'H': 31.29, 'I': 26, 'J': 25.43}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho
    
    # Limpiar bordes filas 2 y 3
    for row in range(2, 4):
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = border_none
    
    # =========================================================================
    # LOGO
    # =========================================================================
    try:
        from PIL import Image as PILImage
        logo_bytes = base64.b64decode(LOGO_BASE64.strip())
        pil_image = PILImage.open(io.BytesIO(logo_bytes))
        
        # Tamaño SICOP: alto 1.39 cm, ancho 7.33 cm
        # Convertir cm a pixeles (aprox 37.8 px/cm)
        width_px = int(7.33 * 37.8)  # ~277 px
        height_px = int(1.39 * 37.8)  # ~53 px
        pil_image = pil_image.resize((width_px, height_px), PILImage.Resampling.LANCZOS)
        
        img_stream = io.BytesIO()
        pil_image.save(img_stream, format="PNG")
        img_stream.seek(0)
        
        logo_img = OpenpyxlImage(img_stream)
        logo_img.width = width_px
        logo_img.height = height_px
        ws.add_image(logo_img, 'A1')
    except Exception as e:
        pass  # Continuar sin logo si hay error
    
    # =========================================================================
    # ENCABEZADO
    # =========================================================================
    ws.merge_cells('A1:J1')
    ws['A1'] = 'Unidad de Administración y Finanzas\nDirección General de Programación, Presupuesto y Finanzas'
    ws['A1'].font = font_title
    ws['A1'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 44.25
    
    # Título
    hoy = date.today()
    ws.merge_cells('A4:J4')
    titulo = f'Estado del ejercicio del 1 de enero al {formatear_fecha(hoy)} por Unidad Responsable de la Secretaría de Agricultura y Desarrollo Rural 1/'
    ws['A4'] = titulo
    ws['A4'].font = font_title
    ws['A4'].alignment = align_center
    ws.row_dimensions[4].height = 22.5
    
    # Fila 5: Vacía
    ws.row_dimensions[5].height = 3.75
    
    # =========================================================================
    # ENCABEZADOS DE COLUMNAS
    # =========================================================================
    headers = [
        'UR', 'Denominación', 'Original\n( a )', 'Modificado anual 2/\n( b )',
        'Modificado al periodo 3/\n( c )', 'Ejercido Acumulado\n(ejercido + devengado + ejercido en trámite)\n( d )',
        'Disponible Anual\n( e ) = ( b ) - ( d )', 'Disponible al periodo\n( f ) = ( c ) - ( d )',
        'Porcentaje de avance anual\n( g ) = ( d ) / ( b )', 'Porcentaje de avance al periodo\n( h ) = ( d ) / ( c )'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_dotted
    ws.row_dimensions[6].height = 106.5
    
    # =========================================================================
    # FUNCIONES PARA ESCRIBIR FILAS
    # =========================================================================
    def escribir_fila_subtotal(fila, texto, datos, es_total=False):
        fill = fill_total if es_total else fill_subtotal
        font_color = font_title if es_total else font_subtotal_white
        
        ws.merge_cells(f'A{fila}:B{fila}')
        cell_texto = ws.cell(row=fila, column=1, value=texto)
        cell_texto.font = font_color
        cell_texto.fill = fill
        cell_texto.alignment = align_right if es_total else align_left
        cell_texto.border = border_dotted
        
        ws.cell(row=fila, column=2).fill = fill
        ws.cell(row=fila, column=2).border = border_dotted
        
        cols_data = ['Original', 'Modificado_anual', 'Modificado_periodo', 'Ejercido_acumulado',
                     'Disponible_anual', 'Disponible_periodo', 'Pct_avance_anual', 'Pct_avance_periodo']
        
        for col_idx, key in enumerate(cols_data, 3):
            cell = ws.cell(row=fila, column=col_idx, value=datos.get(key, 0))
            cell.font = font_color
            cell.fill = fill
            cell.number_format = fmt_pct if col_idx >= 9 else fmt_money
            cell.alignment = Alignment(vertical='top')
            cell.border = border_dotted
        
        ws.row_dimensions[fila].height = 24
    
    def escribir_fila_ur(fila, ur, datos, es_gris=False):
        fill_row = fill_gray if es_gris else fill_white
        denominaciones = config['denominaciones']
        
        cell_ur = ws.cell(row=fila, column=1, value=ur)
        cell_ur.font = font_data
        cell_ur.alignment = align_center
        cell_ur.fill = fill_row
        cell_ur.border = border_dotted
        
        cell_denom = ws.cell(row=fila, column=2, value=denominaciones.get(ur, ''))
        cell_denom.font = font_data
        cell_denom.alignment = align_left
        cell_denom.fill = fill_row
        cell_denom.border = border_dotted
        
        cols_data = ['Original', 'Modificado_anual', 'Modificado_periodo', 'Ejercido_acumulado',
                     'Disponible_anual', 'Disponible_periodo', 'Pct_avance_anual', 'Pct_avance_periodo']
        
        for col_idx, key in enumerate(cols_data, 3):
            cell = ws.cell(row=fila, column=col_idx, value=datos.get(key, 0))
            cell.font = font_data
            cell.fill = fill_row
            cell.number_format = fmt_pct if col_idx >= 9 else fmt_money
            cell.alignment = Alignment(vertical='top')
            cell.border = border_dotted
        
        ws.row_dimensions[fila].height = 24
    
    # =========================================================================
    # ESCRIBIR DATOS
    # =========================================================================
    fila = 7
    contador_fila = 0
    
    # Total general
    escribir_fila_subtotal(fila, 'Total general:', totales, es_total=True)
    fila += 1
    
    # Sector Central
    escribir_fila_subtotal(fila, 'Sector Central', subtotales['sector_central'])
    fila += 1
    contador_fila = 0
    for ur in config['sector_central']:
        datos_ur = resumen[resumen['UR'] == ur]
        if not datos_ur.empty:
            datos = datos_ur.iloc[0].to_dict()
            escribir_fila_ur(fila, ur, datos, es_gris=(contador_fila % 2 == 1))
            fila += 1
            contador_fila += 1
    
    # Oficinas
    escribir_fila_subtotal(fila, 'Oficinas de Representación en las Entidades Federativas', subtotales['oficinas'])
    fila += 1
    contador_fila = 0
    for ur in config['oficinas']:
        datos_ur = resumen[resumen['UR'] == ur]
        if not datos_ur.empty:
            datos = datos_ur.iloc[0].to_dict()
            escribir_fila_ur(fila, ur, datos, es_gris=(contador_fila % 2 == 1))
            fila += 1
            contador_fila += 1
    
    # Órganos Desconcentrados
    escribir_fila_subtotal(fila, 'Órganos Desconcentrados', subtotales['organos_desconcentrados'])
    fila += 1
    contador_fila = 0
    for ur in config['organos_desconcentrados']:
        datos_ur = resumen[resumen['UR'] == ur]
        if not datos_ur.empty:
            datos = datos_ur.iloc[0].to_dict()
            escribir_fila_ur(fila, ur, datos, es_gris=(contador_fila % 2 == 1))
            fila += 1
            contador_fila += 1
    
    # Entidades Paraestatales
    escribir_fila_subtotal(fila, 'Entidades Paraestatales', subtotales['entidades_paraestatales'])
    fila += 1
    contador_fila = 0
    for ur in config['entidades_paraestatales']:
        datos_ur = resumen[resumen['UR'] == ur]
        if not datos_ur.empty:
            datos = datos_ur.iloc[0].to_dict()
            escribir_fila_ur(fila, ur, datos, es_gris=(contador_fila % 2 == 1))
            fila += 1
            contador_fila += 1
    
    # =========================================================================
    # NOTAS AL PIE
    # =========================================================================
    fila += 1
    ultimo_habil = obtener_ultimo_dia_habil(hoy)
    
    ws.merge_cells(f'A{fila}:J{fila}')
    cell_fuente = ws.cell(row=fila, column=1, value=f'Fuente: Elaborado con la base extraída del Sistema de Contabilidad y Presupuesto (SICOP), con corte al {formatear_fecha(ultimo_habil)}.')
    cell_fuente.font = font_notes
    cell_fuente.fill = fill_white
    cell_fuente.border = border_none
    cell_fuente.alignment = align_left
    ws.row_dimensions[fila].height = 35
    fila += 1
    
    ws.merge_cells(f'A{fila}:J{fila}')
    cell_nota1 = ws.cell(row=fila, column=1, value='1/ No Incluye el capítulo 1000 "Servicios personales" ni partida 39801 "Impuesto sobre nóminas".')
    cell_nota1.font = font_notes
    cell_nota1.fill = fill_white
    cell_nota1.border = border_none
    cell_nota1.alignment = align_left
    ws.row_dimensions[fila].height = 35
    fila += 1
    
    ws.merge_cells(f'A{fila}:J{fila}')
    nota_2_texto = f'2/ El Presupuesto Modificado Anual no incluye ${congelados["anual"]:,.2f} ({congelados["texto_anual"]}), recursos congelados.'
    cell_nota2 = ws.cell(row=fila, column=1, value=nota_2_texto)
    cell_nota2.font = font_notes
    cell_nota2.fill = fill_white
    cell_nota2.border = border_none
    cell_nota2.alignment = align_left
    ws.row_dimensions[fila].height = 35
    fila += 1
    
    ws.merge_cells(f'A{fila}:J{fila}')
    nota_3_texto = f'3/ El Presupuesto Modificado al periodo no incluye ${congelados["periodo"]:,.2f} ({congelados["texto_periodo"]}), recursos congelados.'
    cell_nota3 = ws.cell(row=fila, column=1, value=nota_3_texto)
    cell_nota3.font = font_notes
    cell_nota3.fill = fill_white
    cell_nota3.border = border_none
    cell_nota3.alignment = align_left
    ws.row_dimensions[fila].height = 35
    
    # =========================================================================
    # GUARDAR A BYTES
    # =========================================================================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
